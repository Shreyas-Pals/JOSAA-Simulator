"""
Home
"""

from openpyxl import load_workbook
import pandas as pd
import streamlit as slt
Colleges = {}
#For 5 colleges namely, IITB,IITM,IITD,IITK and IITKGP
names = ['B','M','D','K','KGP']

wb = load_workbook('IIT_Preferences_Sample.xlsx')
ws = wb['Sheet1']

#Each college has 5 Branches and each has 5 seats.
for i in names:
    Colleges[f"IIT{i}"] = {
        "Computer Science": 5,
        "Electronics": 5,
        "Data Science": 5,
        "Electrical": 5,
        "Mechanical": 5
}
#Students getting accessed from Students column of the Inputted excel sheet.
class Students:
    def __init__(self,row_number):
        self.name = ws.cell(row=row_number+1, column=2).value
        self.prefs = [ws.cell(row=row_number+1,column=cell).value for cell in range(3,6)]

# Displayed when all seats of colleges have already been taken
failure_message0 = "No College Allotted"
failure_message1 = "No Branch Allotted"
def Allotment(name,colleges,branches):
    # college_result = []
    # branch_result = []
    allocated = False
    for stud_data in zip(colleges,branches): 
        for college_data in list(Colleges.keys()):
            for branch_data in list(Colleges[college_data].keys()):
                seats = Colleges[college_data][branch_data] 
                if (list(stud_data) == [college_data,branch_data] and seats > 0):
                    Student_results = [name,stud_data[0],stud_data[1]]
                    # Student_results.append()
                    # Student_results.append()
                    Colleges[college_data][branch_data] -=1
                    return Student_results
                
    if allocated is False:
        Student_results = [name,failure_message0,failure_message1]
        # Student_results = zip(names,college_result,branch_result)
        return Student_results
                
Student = Students(0)
data = []
i=0
while True:
    i+=1
    Student = Students(i)
    if Student.name == None:
        break

    College_prefs,Branch_prefs = [s.split(" - ")[0] for s in Student.prefs],[s.split(" - ")[1] for s in Student.prefs]
    data.append(Allotment(Student.name,College_prefs,Branch_prefs))

df = pd.DataFrame(data,columns=["Name","College","Branch"])
slt.title(":blue[JOSAA] 2025")

slt.dataframe(df)

# Remaining Seats:
r_seats = []
for college_data in list(Colleges.keys()):
    coll_seats = 0
    for branch_data in list(Colleges[college_data].keys()):
        seats = Colleges[college_data][branch_data]
        coll_seats+=seats
    r_seats.append(coll_seats)

remaining_seats = pd.DataFrame(
    {
    'IITs': list(Colleges.keys()),
    'Seats': r_seats
    }
)

slt.session_state['bar_graph_data'] = remaining_seats