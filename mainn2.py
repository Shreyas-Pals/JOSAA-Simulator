from openpyxl import load_workbook
Colleges = {}
#For 5 colleges namely, IITB,IITM,IITD,IITK and IITKGP
names = ['B','M','D','K','KGP']

wb = load_workbook('IIT_Preferences_Sample.xlsx')
ws = wb['Sheet1']

#Each college has 5 Branches and each has 5 seats.
for i in names:
    Colleges[f"IIT{i}"] = {
        "Computer Science": 5,
        "Electronics": 5,
        "Data Science": 5,
        "Electrical": 5,
        "Mechanical": 5
}
#Students getting accessed from Students column of the Inputted excel sheet.
class Students:
    def __init__(self,row_number):
        self.name = ws.cell(row=row_number+1, column=2).value
        self.prefs = [ws.cell(row=row_number+1,column=cell).value for cell in range(3,6)]

# Displayed when all seats of colleges have already been taken
failure_message = "Sorry to let you know that you haven't gotten a seat in any college"

def Allotment(names,colleges,branches):
    college_result = []
    branch_result = []
    allocated = False
    for stud_data in zip(colleges,branches): 
        for college_data in list(Colleges.keys()):
            for branch_data in list(Colleges[college_data].keys()):
                seats = Colleges[college_data][branch_data] 
                if (list(stud_data) == [college_data,branch_data] and seats > 0):
                    college_result.append(stud_data[0])
                    branch_result.append(stud_data[1])
                    Colleges[college_data][branch_data] -=1
                    Final_results = zip(names,college_result,branch_result)
                    allocated = True
                    return list((Final_results))
                
    if allocated is False:
        college_result.append(failure_message)
        branch_result.append("")
        Final_results = zip(names,college_result,branch_result)
        return list((Final_results))
                
Student = Students(0)
i=0
while True:
    Student_names = []
    i+=1
    Student = Students(i)
    if Student.name == None:
        break
    else:
        Student_names.append(Student.name)

    College_prefs,Branch_prefs = [s.split(" - ")[0] for s in Student.prefs],[s.split(" - ")[1] for s in Student.prefs]
    print(Allotment(Student_names,College_prefs,Branch_prefs))
